'''
IMDb scraper
'''

import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
import json
import re


DB_PATH = "MovieGenreIGC_v3.xlsx"
OUTPUT = "output3.json"

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(DB_PATH)
 
# Define variable to read sheet
dataframe1 = dataframe.active
    

HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:47.0) Gecko/20100101 Firefox/102.0"}



def getData(url: str) -> dict:
    '''
    Returns a dictionary with title, director, description, year, and genres
    of the movie specified in the IMDb url.
    Returns an empty dictionary if the page doesn't exist.
    '''

    if "http://www.imdb.com/title/tt" not in url:
        return {}

    r = requests.get(url, headers = HEADERS)

    # print(r.status_code, url)

    if r.status_code != 200:
        return {}


    imdb = BeautifulSoup(r.text, "html.parser")

    # check if gay
    if ("gay" in r.text or "lgbt" in r.text):
        gay = True
    else:
        gay = False

    # movie title
    title = imdb.select_one("#__next > main > div > section.ipc-page-background.ipc-page-background--base.sc-9b716f3b-0.hWwhTB > section > div:nth-child(4) > section > section > div.sc-80d4314-0.fjPRnj > div.sc-80d4314-1.fbQftq > h1")
    # director
    director = imdb.select_one("#__next > main > div > section.ipc-page-background.ipc-page-background--base.sc-9b716f3b-0.hWwhTB > section > div:nth-child(4) > section > section > div.sc-7643a8e3-6.bunqBa > div.sc-7643a8e3-10.itwFpV > div.sc-7643a8e3-8.hOtbMQ > div.sc-fa02f843-2.dwQKsL > div > div > ul > li:nth-child(1) > div > ul > li > a")
    # description
    desc = imdb.select_one("#__next > main > div > section.ipc-page-background.ipc-page-background--base.sc-9b716f3b-0.hWwhTB > section > div:nth-child(4) > section > section > div.sc-7643a8e3-6.bunqBa > div.sc-7643a8e3-10.itwFpV > div.sc-7643a8e3-8.hOtbMQ > div.sc-16ede01-9.bbiYSi.sc-7643a8e3-11.efPxUc > div.sc-16ede01-7.hrgVKw > span.sc-16ede01-2.gXUyNh")
    # year
    year = imdb.select_one("#__next > main > div > section.ipc-page-background.ipc-page-background--base.sc-9b716f3b-0.hWwhTB > section > div:nth-child(4) > section > section > div.sc-80d4314-0.fjPRnj > div.sc-80d4314-1.fbQftq > div > ul > li:nth-child(1) > span")


    # genres
    genres = imdb.select("#__next > main > div > section.ipc-page-background.ipc-page-background--base.sc-9b716f3b-0.hWwhTB > section > div:nth-child(4) > section > section > div.sc-7643a8e3-6.bunqBa > div.sc-7643a8e3-10.itwFpV > div.sc-7643a8e3-8.hOtbMQ > div.sc-16ede01-9.bbiYSi.sc-7643a8e3-11.efPxUc > div.ipc-chip-list--baseAlt.ipc-chip-list.sc-16ede01-5.ggbGKe > div.ipc-chip-list__scroller > a > span")
    parsed_genres = []

    for elem in genres:
        parsed_genres.append(elem.text)

    try:
        return {
            "title": title.text,
            "director": director.text,
            "year": year.text,
            "description": desc.text,
            "genres": parsed_genres,
            "lgbt": gay
        }

    except AttributeError:
        return {}


def main():
    '''
    Goes through the database on DB_PATH and gets the urls,
    using those to scrape the data of each movie, and saving it to OUTPUT
    '''
    data = {}
    data['movies'] = []

    result = dict()
    idCount = 0
    # Iterate the loop to read the excel cell values 
    # containing the url to acces each movie iMDB url
    with open(OUTPUT, 'w') as file:
        for row in range(1, 4 ):  #dataframe1.max_row
            for col in dataframe1.iter_cols(2, 2):
                #print(col[row].value)
                movieUrl = col[row].value
                result = getData(movieUrl)

                if result != {}:
                    
                    json.dump({'index': {'_id': str(idCount)}}, file)
                    file.write("\n")

                    json.dump({
                        'movie_title': result["title"],
                        'movie_director': result["director"],
                        'movie_year': result["year"],
                        'movie_description': result["description"],
                        'movie_genres': result['genres'],
                        "movie-lgbt": result["lgbt"]
                        }, file)

                    file.write("\n")
                
                idCount += 1

            
    #with open(OUTPUT, 'w') as file:
        #json.dump(data, file)

    pass


def test():
    # type 1
    print(getData("http://www.imdb.com/title/tt0357608"))  # one genre
    print(getData("http://www.imdb.com/title/tt4991286"))  # multiple genres
    print(getData("http://www.imdb.com/title/tt5726616"))  # gay, but not right format
    print(getData("http://www.imdb.com/title/tt0adfdf7608"))  # fails


if __name__ == "__main__":
    # test()
    main()
